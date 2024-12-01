#멜론에서 장르와 하트 수 가져와서 엑셀에 저장 코드
import openpyxl
import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time

#파일 경로 설정
Location = 'C:/Users/SAMSUNG/Desktop/해커톤/Song List.xlsx'
wb = openpyxl.load_workbook(Location, data_only=True)

sheet_list = wb.get_sheet_names()

w1 = wb[sheet_list[0]]
#크롬 키기
driver = webdriver.Chrome("./chromedriver")

#멜론 사이트 들어가기
driver.get('http://www.melon.com')
time.sleep(2)

#엑셀에 값 받아와 새로운 엑셀 파일에 저장 (try/execpt문으로 예외 발생시 바로 저장) 
try:
    for i in range(2, w1.max_row):
        #몇번째 도는지 확인
        print(i)
        # 값 받아오기
        a = w1.cell(row = i, column = 3).value

        #검색창 클릭
        search = driver.find_element(By.CSS_SELECTOR, 'input.ui-autocomplete-input')
        search.click()

        #검색어 입력
        search.send_keys(a)
        search.send_keys(Keys.ENTER)
        time.sleep(1)

        #곡 정보 들어가기(예외처리)
        try:
            driver.find_element("xpath",'//*[@id="frm_searchSong"]/div/table/tbody/tr[1]/td[3]/div/div/a[1]/span').click()
        except: #하나만 검색될경우
            try: 
                driver.find_element("xpath",'//*[@id="frm_songList"]/div/table/tbody/tr/td[3]/div/div/a[1]/span').click()
            except: #검색안될 경우(시작 페이지로)
                driver.get('http://www.melon.com')
                time.sleep(2)
                continue
        time.sleep(1)

        #곡 장르 불러오기
        element = driver.find_element("xpath", '//*[@id="downloadfrm"]/div/div/div[2]/div[2]/dl/dd[3]').text
        
        #곡 하트 불러오기
        heart = driver.find_element("xpath", '//*[@id="d_like_count"]').text
        
        #값 저장하기
        w1.cell(row = i, column = 6).value = element
        w1.cell(row = i, column = 7).value = heart
        time.sleep(1)
except:
    #예외 발생시 바로 저장하고 종료 
    new_file='C:/Users/SAMSUNG/Desktop/해커톤/Song_List2.xlsx'
    wb.save(new_file)

# 파일 저장    
new_file='C:/Users/SAMSUNG/Desktop/해커톤/Song_List2.xlsx'
wb.save(new_file)

