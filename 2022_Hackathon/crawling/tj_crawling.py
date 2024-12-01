#멜론(음원 서비스)에서 원하는 곡의 장르 및 좋아요 수 크롤링해서 엑셀에 저장해주는 코드

import openpyxl
import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time

#파일 경로 설정
Location = 'C:/Users/SAMSUNG/Desktop/해커톤/Song List.xlsx'
wb = openpyxl.load_workbook(Location, data_only=True)

sheet_list = wb.get_sheet_names()

w1 = wb[sheet_list[0]]

#크롬 켜기
driver = webdriver.Chrome("./chromedriver")

#멜론 사이트 들어가기
driver.get('https://www.tjmedia.com/tjsong/song_search_list.asp')
time.sleep(2)
try:
    for i in range(2, w1.max_row):
        print(i)
        a = w1.cell(row = i, column = 3).value

        driver.find_element(By.CSS_SELECTOR, 'select.searchSelectSongForm1').click()
        driver.find_element("xpath",'//*[@id="searchSong"]/form/ul[1]/li/select[1]/option[2]').click()
        driver.find_element(By.CSS_SELECTOR, 'select.searchSelectSongForm2').click()
        driver.find_element("xpath",'//*[@id="searchSong"]/form/ul[1]/li/select[2]/option[2]').click()

        search = driver.find_element(By.CSS_SELECTOR, 'input.search_input2.sort_t')
        search.click()
        
        search.send_keys(a)
        search.send_keys(Keys.ENTER)
        time.sleep(1)

        try:
            s = driver.find_element("xpath",'//*[@id="BoardType1"]/table/tbody/tr[2]/td[1]').text
            search = driver.find_element(By.CSS_SELECTOR, 'input.search_input2.sort_t')
            search.click()
            search.send_keys(Keys.CONTROL, 'a')
            search.send_keys(Keys.BACKSPACE)
        except:
            search = driver.find_element(By.CSS_SELECTOR, 'input.search_input2.sort_t')
            search.click()
            search.send_keys(Keys.CONTROL, 'a')
            search.send_keys(Keys.BACKSPACE)
            #continue
        #값 저장하기
        w1.cell(row = i, column = 8).value = s
        time.sleep(3) #잦은 접속으로 인한 튕귐 우회
except:
    new_file = "C:/Users/SAMSUNG/Desktop/해커톤/Song List2.xlsx"
    wb.save(new_file)
new_file = "C:/Users/SAMSUNG/Desktop/해커톤/Song List2.xlsx"
wb.save(new_file)